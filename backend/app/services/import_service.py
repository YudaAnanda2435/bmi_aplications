from collections import Counter
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.models.classification import ClassificationResult
from app.models.resident import Resident
from app.schemas.common import FREQUENCY_VALUES, GENDER_VALUES, TRANSPORT_VALUES, YES_NO_VALUES
from app.services.ml_service import MLService
from app.utils.measurement import normalize_height_to_meter


REQUIRED_COLUMNS = [
    "name",
    "gender",
    "age",
    "height",
    "weight",
    "family_history_with_overweight",
    "favc",
    "fcvc",
    "ncp",
    "caec",
    "smoke",
    "ch2o",
    "scc",
    "faf",
    "tue",
    "calc",
    "mtrans",
]

INPUT_SHEET_NAME = "Input_Penduduk"
GUIDE_SHEET_NAME = "Panduan_Pengisian"
CODES_SHEET_NAME = "Kode_Sistem"
TECHNICAL_SHEET_NAME = "resident_import"

INDONESIAN_TO_BACKEND_COLUMNS = {
    "Nama Penduduk": "name",
    "Jenis Kelamin": "gender",
    "Usia": "age",
    "Tinggi Badan (cm)": "height",
    "Berat Badan (kg)": "weight",
    "Riwayat Keluarga Obesitas": "family_history_with_overweight",
    "Konsumsi Makanan Tinggi Kalori / Fast Food": "favc",
    "Frekuensi Konsumsi Sayur": "fcvc",
    "Jumlah Makan Utama": "ncp",
    "Kebiasaan Makan di Antara Waktu Makan": "caec",
    "Merokok": "smoke",
    "Konsumsi Air Harian": "ch2o",
    "Kebiasaan Mengontrol Kalori": "scc",
    "Frekuensi Aktivitas Fisik": "faf",
    "Durasi Penggunaan Perangkat Teknologi": "tue",
    "Konsumsi Alkohol": "calc",
    "Jenis Transportasi Harian": "mtrans",
}
INDONESIAN_COLUMNS = list(INDONESIAN_TO_BACKEND_COLUMNS.keys())
BACKEND_TO_INDONESIAN_COLUMNS = {
    backend_column: indonesia_column
    for indonesia_column, backend_column in INDONESIAN_TO_BACKEND_COLUMNS.items()
}

YES_NO_FIELDS = {"family_history_with_overweight", "favc", "smoke", "scc"}
FREQUENCY_FIELDS = {"caec", "calc"}
NUMERIC_LIFESTYLE_FIELDS = {"fcvc", "ncp", "ch2o", "faf", "tue"}

INDONESIAN_GENDER_VALUES = {
    "Laki-laki": "Male",
    "Perempuan": "Female",
}
INDONESIAN_YES_NO_VALUES = {
    "Ya": "yes",
    "Tidak": "no",
}
INDONESIAN_FCVC_VALUES = {
    "Jarang": 1,
    "Kadang-kadang": 2,
    "Sering": 3,
}
INDONESIAN_NCP_VALUES = {
    "1 kali": 1,
    "2 kali": 2,
    "3 kali": 3,
    "4 kali": 4,
    "Lebih dari 3 kali": 4,
}
INDONESIAN_FREQUENCY_VALUES = {
    "Tidak pernah": "no",
    "Kadang-kadang": "Sometimes",
    "Sering": "Frequently",
    "Selalu": "Always",
}
INDONESIAN_CH2O_VALUES = {
    "Rendah": 1,
    "Sedang": 2,
    "Tinggi": 3,
}
INDONESIAN_FAF_VALUES = {
    "Tidak pernah": 0,
    "Rendah": 1,
    "Sedang": 2,
    "Tinggi": 3,
}
INDONESIAN_TUE_VALUES = {
    "Rendah": 0,
    "Sedang": 1,
    "Tinggi": 2,
}
INDONESIAN_MTRANS_VALUES = {
    "Mobil": "Automobile",
    "Motor": "Motorbike",
    "Sepeda": "Bike",
    "Transportasi Umum": "Public_Transportation",
    "Jalan Kaki": "Walking",
}

INDONESIAN_VALUE_MAPPINGS = {
    "gender": INDONESIAN_GENDER_VALUES,
    "family_history_with_overweight": INDONESIAN_YES_NO_VALUES,
    "favc": INDONESIAN_YES_NO_VALUES,
    "smoke": INDONESIAN_YES_NO_VALUES,
    "scc": INDONESIAN_YES_NO_VALUES,
    "fcvc": INDONESIAN_FCVC_VALUES,
    "ncp": INDONESIAN_NCP_VALUES,
    "caec": INDONESIAN_FREQUENCY_VALUES,
    "ch2o": INDONESIAN_CH2O_VALUES,
    "faf": INDONESIAN_FAF_VALUES,
    "tue": INDONESIAN_TUE_VALUES,
    "calc": INDONESIAN_FREQUENCY_VALUES,
    "mtrans": INDONESIAN_MTRANS_VALUES,
}

SAMPLE_ROW = {
    "name": "Budi Pratama",
    "gender": "Male",
    "age": 26,
    "height": 160,
    "weight": 105,
    "family_history_with_overweight": "yes",
    "favc": "yes",
    "fcvc": 3,
    "ncp": 3,
    "caec": "Frequently",
    "smoke": "no",
    "ch2o": 3,
    "scc": "no",
    "faf": 2,
    "tue": 2,
    "calc": "Sometimes",
    "mtrans": "Public_Transportation",
}

INDONESIAN_SAMPLE_ROW = {
    "Nama Penduduk": "Budi Pratama",
    "Jenis Kelamin": "Laki-laki",
    "Usia": 26,
    "Tinggi Badan (cm)": 160,
    "Berat Badan (kg)": 105,
    "Riwayat Keluarga Obesitas": "Ya",
    "Konsumsi Makanan Tinggi Kalori / Fast Food": "Ya",
    "Frekuensi Konsumsi Sayur": "Sering",
    "Jumlah Makan Utama": "3 kali",
    "Kebiasaan Makan di Antara Waktu Makan": "Sering",
    "Merokok": "Tidak",
    "Konsumsi Air Harian": "Tinggi",
    "Kebiasaan Mengontrol Kalori": "Tidak",
    "Frekuensi Aktivitas Fisik": "Sedang",
    "Durasi Penggunaan Perangkat Teknologi": "Tinggi",
    "Konsumsi Alkohol": "Kadang-kadang",
    "Jenis Transportasi Harian": "Transportasi Umum",
}


def create_import_template() -> BytesIO:
    workbook = Workbook()
    input_sheet = workbook.active
    input_sheet.title = INPUT_SHEET_NAME
    input_sheet.append(INDONESIAN_COLUMNS)
    input_sheet.append([INDONESIAN_SAMPLE_ROW[column] for column in INDONESIAN_COLUMNS])
    input_sheet.freeze_panes = "A2"

    for cell in input_sheet[1]:
        cell.style = "Headline 3"

    _set_column_widths(input_sheet, INDONESIAN_COLUMNS)
    _add_input_dropdowns(input_sheet)

    guide_sheet = workbook.create_sheet(GUIDE_SHEET_NAME)
    guide_sheet.append(["Panduan", "Keterangan"])
    guide_rows = [
        ("Tujuan file", "Template ini digunakan untuk import data penduduk dan klasifikasi massal."),
        ("Sheet yang diisi", "Isi data pada sheet Input_Penduduk."),
        ("BMI", "Tidak perlu diisi. Backend menghitung BMI otomatis dari tinggi dan berat badan."),
        ("Tinggi badan", "Boleh diisi cm seperti 160. Backend menyimpan dan memproses dalam meter."),
        ("Baris kosong", "Baris yang benar-benar kosong akan diabaikan saat import."),
        ("Training model", "File ini bukan untuk upload dataset training."),
    ]
    for row in guide_rows:
        guide_sheet.append(row)
    for cell in guide_sheet[1]:
        cell.style = "Headline 3"
    _set_column_widths(guide_sheet, ["Panduan", "Keterangan"])

    codes_sheet = workbook.create_sheet(CODES_SHEET_NAME)
    codes_sheet.append(["Field", "Pilihan Petugas", "Nilai Sistem"])
    _append_code_rows(codes_sheet)
    for cell in codes_sheet[1]:
        cell.style = "Headline 3"
    _set_column_widths(codes_sheet, ["Field", "Pilihan Petugas", "Nilai Sistem"])

    technical_sheet = workbook.create_sheet(TECHNICAL_SHEET_NAME)
    technical_sheet.append(REQUIRED_COLUMNS)
    technical_sheet.append([SAMPLE_ROW[column] for column in REQUIRED_COLUMNS])
    technical_sheet.freeze_panes = "A2"
    for cell in technical_sheet[1]:
        cell.style = "Headline 3"
    _set_column_widths(technical_sheet, REQUIRED_COLUMNS)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def parse_and_validate_excel(file_bytes: bytes) -> dict[str, Any]:
    rows, header_errors = _read_excel_rows(file_bytes)
    valid_rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = header_errors

    if header_errors:
        return {
            "total_rows": len(rows),
            "valid_count": 0,
            "error_count": len(errors),
            "valid_rows": [],
            "preview_valid_data": [],
            "errors": errors,
        }

    for row in rows:
        row_errors: list[dict[str, Any]] = []
        normalized = _validate_row(
            row["row_number"],
            row["data"],
            row_errors,
            row["field_labels"],
            row["source"],
        )

        if row_errors:
            errors.extend(row_errors)
            continue

        valid_rows.append(normalized)

    preview_valid_data = [_serialize_valid_row(row) for row in valid_rows]

    return {
        "total_rows": len(rows),
        "valid_count": len(valid_rows),
        "error_count": len(errors),
        "valid_rows": valid_rows,
        "preview_valid_data": preview_valid_data,
        "errors": errors,
    }


def import_and_classify_residents(
    db: Session,
    file_bytes: bytes,
    created_by: int | None,
    ml_service: MLService,
) -> dict[str, Any]:
    validation_result = parse_and_validate_excel(file_bytes)
    errors = list(validation_result["errors"])
    class_distribution: Counter[str] = Counter()
    warning_distribution: Counter[str] = Counter()
    imported_count = 0

    for row in validation_result["valid_rows"]:
        try:
            resident = Resident(created_by=created_by, **_resident_fields(row))
            db.add(resident)
            db.flush()

            prediction = ml_service.predict(_model_features(row))
            probabilities = prediction["probabilities"]
            classification_result = ClassificationResult(
                resident_id=resident.id,
                predicted_class=prediction["predicted_class"],
                probability_underweight=probabilities.get("Underweight"),
                probability_normal=probabilities.get("Normal"),
                probability_overweight=probabilities.get("Overweight"),
                probability_obesity=probabilities.get("Obesity"),
                recommendation=prediction["recommendation"],
                early_warning=prediction["early_warning"],
                note=prediction["note"],
            )
            db.add(classification_result)
            db.commit()

            imported_count += 1
            class_distribution[prediction["predicted_class"]] += 1
            warning_distribution[prediction["early_warning"]] += 1
        except Exception as exc:
            db.rollback()
            errors.append(
                {
                    "row_number": row["row_number"],
                    "field": "row",
                    "message": f"Failed to import and classify row: {exc}",
                }
            )

    return {
        "total_rows": validation_result["total_rows"],
        "imported_count": imported_count,
        "failed_count": validation_result["total_rows"] - imported_count,
        "class_distribution": dict(class_distribution),
        "warning_distribution": dict(warning_distribution),
        "errors": errors,
    }


def _read_excel_rows(file_bytes: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    errors: list[dict[str, Any]] = []

    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        return [], [{"row_number": 1, "field": "file", "message": f"Invalid Excel file: {exc}"}]

    if INPUT_SHEET_NAME in workbook.sheetnames:
        return _read_indonesian_sheet(workbook[INPUT_SHEET_NAME])

    worksheet = workbook[TECHNICAL_SHEET_NAME] if TECHNICAL_SHEET_NAME in workbook.sheetnames else workbook.active
    header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [_clean_header(value) for value in header_values or []]

    missing_columns = [column for column in REQUIRED_COLUMNS if column not in headers]
    for column in missing_columns:
        errors.append(
            {
                "row_number": 1,
                "field": column,
                "message": "Required column is missing",
            }
        )

    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if _is_empty_row(values):
            continue

        row_data = {
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
            if header
        }
        rows.append(
            {
                "row_number": row_number,
                "data": row_data,
                "field_labels": {column: column for column in REQUIRED_COLUMNS},
                "source": "technical",
            }
        )

    return rows, errors


def _read_indonesian_sheet(worksheet: Any) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    errors: list[dict[str, Any]] = []
    header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [_clean_header(value) for value in header_values or []]

    missing_columns = [column for column in INDONESIAN_COLUMNS if column not in headers]
    for column in missing_columns:
        errors.append(
            {
                "row_number": 1,
                "field": column,
                "message": "Kolom wajib tidak ditemukan",
            }
        )

    rows: list[dict[str, Any]] = []
    field_labels = BACKEND_TO_INDONESIAN_COLUMNS.copy()

    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if _is_empty_row(values):
            continue

        raw_row_data = {
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
            if header
        }
        row_data = {
            backend_column: raw_row_data.get(indonesia_column)
            for indonesia_column, backend_column in INDONESIAN_TO_BACKEND_COLUMNS.items()
        }
        rows.append(
            {
                "row_number": row_number,
                "data": row_data,
                "field_labels": field_labels,
                "source": "indonesian",
            }
        )

    return rows, errors


def _validate_row(
    row_number: int,
    row_data: dict[str, Any],
    errors: list[dict[str, Any]],
    field_labels: dict[str, str],
    source: str,
) -> dict[str, Any]:
    name = _parse_required_text(row_number, _label("name", field_labels), row_data.get("name"), errors)
    gender = _parse_mapped_or_category(
        row_number,
        "gender",
        row_data.get("gender"),
        GENDER_VALUES,
        errors,
        field_labels,
        source,
    )
    age = _parse_positive_int(row_number, _label("age", field_labels), row_data.get("age"), errors)
    height_raw = _parse_positive_float(row_number, _label("height", field_labels), row_data.get("height"), errors)
    weight = _parse_positive_float(row_number, _label("weight", field_labels), row_data.get("weight"), errors)

    normalized: dict[str, Any] = {
        "row_number": row_number,
        "name": name,
        "gender": gender,
        "age": age,
        "height": None,
        "weight": weight,
        "family_history_with_overweight": None,
        "favc": None,
        "fcvc": None,
        "ncp": None,
        "caec": None,
        "smoke": None,
        "ch2o": None,
        "scc": None,
        "faf": None,
        "tue": None,
        "calc": None,
        "mtrans": None,
        "bmi": None,
    }

    if height_raw is not None:
        normalized["height"] = normalize_height_to_meter(height_raw)

    for field in YES_NO_FIELDS:
        normalized[field] = _parse_mapped_or_category(
            row_number,
            field,
            row_data.get(field),
            YES_NO_VALUES,
            errors,
            field_labels,
            source,
        )

    for field in FREQUENCY_FIELDS:
        normalized[field] = _parse_mapped_or_category(
            row_number,
            field,
            row_data.get(field),
            FREQUENCY_VALUES,
            errors,
            field_labels,
            source,
        )

    normalized["mtrans"] = _parse_mapped_or_category(
        row_number,
        "mtrans",
        row_data.get("mtrans"),
        TRANSPORT_VALUES,
        errors,
        field_labels,
        source,
    )

    for field in NUMERIC_LIFESTYLE_FIELDS:
        normalized[field] = _parse_mapped_or_float(
            row_number,
            field,
            row_data.get(field),
            errors,
            field_labels,
            source,
        )

    if normalized["height"] is not None and weight is not None:
        normalized["bmi"] = round(weight / (normalized["height"] ** 2), 2)

    return normalized


def _resident_fields(row: dict[str, Any]) -> dict[str, Any]:
    return {column: row[column] for column in REQUIRED_COLUMNS} | {"bmi": row["bmi"]}


def _model_features(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "Gender": row["gender"],
        "Age": row["age"],
        "Height": row["height"],
        "Weight": row["weight"],
        "family_history_with_overweight": row["family_history_with_overweight"],
        "FAVC": row["favc"],
        "FCVC": row["fcvc"],
        "NCP": row["ncp"],
        "CAEC": row["caec"],
        "SMOKE": row["smoke"],
        "CH2O": row["ch2o"],
        "SCC": row["scc"],
        "FAF": row["faf"],
        "TUE": row["tue"],
        "CALC": row["calc"],
        "MTRANS": row["mtrans"],
    }


def _serialize_valid_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        column: row[column]
        for column in REQUIRED_COLUMNS
    } | {"bmi": row["bmi"]}


def _clean_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _label(field: str, field_labels: dict[str, str]) -> str:
    return field_labels.get(field, field)


def _is_empty_row(values: tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _parse_required_text(
    row_number: int,
    field: str,
    value: Any,
    errors: list[dict[str, Any]],
) -> str | None:
    if _is_blank(value):
        errors.append({"row_number": row_number, "field": field, "message": "Field is required"})
        return None
    return str(value).strip()


def _parse_category(
    row_number: int,
    field: str,
    value: Any,
    allowed_values: list[str],
    errors: list[dict[str, Any]],
) -> str | None:
    text_value = _parse_required_text(row_number, field, value, errors)
    if text_value is None:
        return None
    if text_value not in allowed_values:
        errors.append(
            {
                "row_number": row_number,
                "field": field,
                "message": f"Value must be one of: {', '.join(allowed_values)}",
            }
        )
        return None
    return text_value


def _parse_mapped_or_category(
    row_number: int,
    field: str,
    value: Any,
    allowed_values: list[str],
    errors: list[dict[str, Any]],
    field_labels: dict[str, str],
    source: str,
) -> str | None:
    if source != "indonesian":
        return _parse_category(row_number, field, value, allowed_values, errors)

    label = _label(field, field_labels)
    text_value = _parse_required_text(row_number, label, value, errors)
    if text_value is None:
        return None

    value_map = INDONESIAN_VALUE_MAPPINGS[field]
    if text_value not in value_map:
        errors.append(
            {
                "row_number": row_number,
                "field": label,
                "message": f"Nilai harus salah satu dari: {', '.join(value_map.keys())}",
            }
        )
        return None

    return value_map[text_value]


def _parse_mapped_or_float(
    row_number: int,
    field: str,
    value: Any,
    errors: list[dict[str, Any]],
    field_labels: dict[str, str],
    source: str,
) -> float | None:
    if source != "indonesian":
        return _parse_float(row_number, field, value, errors)

    label = _label(field, field_labels)
    text_value = _parse_required_text(row_number, label, value, errors)
    if text_value is None:
        return None

    value_map = INDONESIAN_VALUE_MAPPINGS[field]
    if text_value not in value_map:
        errors.append(
            {
                "row_number": row_number,
                "field": label,
                "message": f"Nilai harus salah satu dari: {', '.join(value_map.keys())}",
            }
        )
        return None

    return float(value_map[text_value])


def _parse_float(
    row_number: int,
    field: str,
    value: Any,
    errors: list[dict[str, Any]],
) -> float | None:
    if _is_blank(value) or isinstance(value, bool):
        errors.append({"row_number": row_number, "field": field, "message": "Value must be a number"})
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        errors.append({"row_number": row_number, "field": field, "message": "Value must be a number"})
        return None


def _parse_positive_float(
    row_number: int,
    field: str,
    value: Any,
    errors: list[dict[str, Any]],
) -> float | None:
    parsed_value = _parse_float(row_number, field, value, errors)
    if parsed_value is not None and parsed_value <= 0:
        errors.append({"row_number": row_number, "field": field, "message": "Value must be greater than 0"})
        return None
    return parsed_value


def _parse_positive_int(
    row_number: int,
    field: str,
    value: Any,
    errors: list[dict[str, Any]],
) -> int | None:
    parsed_value = _parse_positive_float(row_number, field, value, errors)
    if parsed_value is None:
        return None
    if parsed_value % 1 != 0:
        errors.append({"row_number": row_number, "field": field, "message": "Value must be a whole number"})
        return None
    return int(parsed_value)


def _set_column_widths(worksheet: Any, headers: list[str]) -> None:
    for index, header in enumerate(headers, start=1):
        width = max(14, min(len(header) + 4, 42))
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = width


def _add_input_dropdowns(worksheet: Any) -> None:
    dropdowns = {
        "B": list(INDONESIAN_GENDER_VALUES.keys()),
        "F": list(INDONESIAN_YES_NO_VALUES.keys()),
        "G": list(INDONESIAN_YES_NO_VALUES.keys()),
        "H": list(INDONESIAN_FCVC_VALUES.keys()),
        "I": list(INDONESIAN_NCP_VALUES.keys()),
        "J": list(INDONESIAN_FREQUENCY_VALUES.keys()),
        "K": list(INDONESIAN_YES_NO_VALUES.keys()),
        "L": list(INDONESIAN_CH2O_VALUES.keys()),
        "M": list(INDONESIAN_YES_NO_VALUES.keys()),
        "N": list(INDONESIAN_FAF_VALUES.keys()),
        "O": list(INDONESIAN_TUE_VALUES.keys()),
        "P": list(INDONESIAN_FREQUENCY_VALUES.keys()),
        "Q": list(INDONESIAN_MTRANS_VALUES.keys()),
    }

    for column, options in dropdowns.items():
        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(options)}"',
            allow_blank=False,
        )
        worksheet.add_data_validation(validation)
        validation.add(f"{column}2:{column}500")


def _append_code_rows(worksheet: Any) -> None:
    code_groups = {
        "Jenis Kelamin": INDONESIAN_GENDER_VALUES,
        "Ya/Tidak": INDONESIAN_YES_NO_VALUES,
        "Frekuensi Konsumsi Sayur": INDONESIAN_FCVC_VALUES,
        "Jumlah Makan Utama": INDONESIAN_NCP_VALUES,
        "Kebiasaan Makan di Antara Waktu Makan": INDONESIAN_FREQUENCY_VALUES,
        "Konsumsi Air Harian": INDONESIAN_CH2O_VALUES,
        "Frekuensi Aktivitas Fisik": INDONESIAN_FAF_VALUES,
        "Durasi Penggunaan Perangkat Teknologi": INDONESIAN_TUE_VALUES,
        "Konsumsi Alkohol": INDONESIAN_FREQUENCY_VALUES,
        "Jenis Transportasi Harian": INDONESIAN_MTRANS_VALUES,
    }
    for field, options in code_groups.items():
        for display_value, system_value in options.items():
            worksheet.append([field, display_value, system_value])
